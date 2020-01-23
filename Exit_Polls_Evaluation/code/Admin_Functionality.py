import openpyxl as op

path = r"D:\KTyagi\KT\Python_Workspace\DataScience\Exit Polls Mini Project\Exit_Poll_Input.xlsx"
wk_book = op.load_workbook(path)
sheet_obj = wk_book.active

def create_header():
	tuple1 = ('SR','CANDIDATE','PARTY','CRIMINAL_RECORD','SEVERITY','FAMILY_MEMBERS IN POLITICS','PROPERTY')
	sheet_obj.append(tuple1)
	wk_book.save(path)

def add_candidate():
	for i in range(100):
		for j in range(2,102):
			for k in range(1, sheet_obj.max_column +1):
				sr = i+1
				name = raw_input('Name Of the candidate: ')
				party = raw_input('Name Of the party: ')
				criminal_rec =raw_input('Criminal Record (Y/N)').upper()
				if criminal_rec == 'Y' or criminal_rec == 'YES':
					severity = int(raw_input('What is the severity (1 - 9): '))
				else:
					severity = 0

				fam_members = int(raw_input('How many family members in politics: '))
				property_amount = int(raw_input('How much property the candidate holds?'))

				t1 = (sr,name,party,criminal_rec,severity,fam_members,property_amount)
				sheet_obj.append(t1)
				wk_book.save(path)

def update_candidate():
	sr_no = int(raw_input('Which number you want to update the details'))
	for j in range(2,sheet_obj.max_row+1):
			if ((j-1) == sr_no):
				while True:
					print 'What details you want to update: '
					print '\n\t 1. NAME \n\t 2. PARTY \n\t 3. CRIMINAL_RECORD \n\t 4. SEVERITY \n\t 5. FAMILY_MEMBERS \n\t 6. PROPERTY \n\t 7. EXIT'
					ch = int(raw_input('Your choice: '))
					if ch == 1:
						cell_obj = sheet_obj.cell(row = j,column = 2)
						cell_obj.value = raw_input('Enter the updated name: ').upper()
					elif ch == 2:
						cell_obj = sheet_obj.cell(row = j,column = 3)
						cell_obj.value = raw_input('Enter the updated party name: ').upper()
					elif ch == 3:
						cell_obj = sheet_obj.cell(row = j,column = 4)
						cell_obj.value = raw_input('Enter the updated criminal record: ').upper()
					elif ch == 4:
						if(sheet_obj.cell(row = j,column =4) == 'Y' or sheet_obj.cell(row = j,column =4) == 'YES'):
							cell_obj = sheet_obj.cell(row = j,column = 5)
							cell_obj.value = int(raw_input('Enter the updated severity: '))
						else:
							cell_obj = sheet_obj.cell(row = j,column = 5)
							cell_obj.value = 0
					elif ch == 5:
						cell_obj = sheet_obj.cell(row = j,column = 6)
						cell_obj.value = int(raw_input('Enter the updated count of candidate\'s family members in politics: '))
					elif ch == 6:
						cell_obj = sheet_obj.cell(row = j,column = 7)
						cell_obj.value = int(raw_input('Enter the updated amount of candidate\'s property: '))
					elif ch == 7:
						break
					else:
						print 'Wrong Choice....'
					wk_book.save(path)
			else:
				pass

def delete_candidate():
	sr_no = int(raw_input('Which candidate record you want to remove from the list: '))
	for j in range(2,sheet_obj.max_row+1):
			if ((j-1) == sr_no):
				sheet_obj.delete_rows(sr_no +1)
				wk_book.save(path)
			else:
				pass

def main():
	print '*' * 80
	print '\n\t\t\t...WELCOME TO THE GAME OF POLITICS....'
	print '*' * 80
	create_header()
	while True:
		print '\n \t\t\t MENU\n \t\t\t 1. ADD CANDIDATE\n \t\t\t 2. UPDATE CANDIDATE\n \t\t\t 3. DELETE CANDIDATE'
		ch = int(raw_input('Enter your choice: '))
		if ch == 1:
			add_candidate()
		elif ch == 2:
			update_candidate()
		elif ch == 3:
			delete_candidate()
		else:
			print 'WRONG CHOICE...'
			break

main()