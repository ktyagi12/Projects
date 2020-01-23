import openpyxl as op
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sb

path = r"D:\KTyagi\KT\Python_Workspace\DataScience\Exit Polls Mini Project\Exit_Poll_Input.xlsx"
path1 = r"D:\KTyagi\KT\Python_Workspace\DataScience\Exit Polls Mini Project\Exit_Poll_Results.xlsx"


workbook = op.load_workbook(path)
workbook1 = op.load_workbook(path1)
sheet_obj = workbook.active
sheet_obj1 = workbook1.active

input_data = {}
result_data = {}
temp = {}

for i in range(2, sheet_obj.max_row +1):
	sr_no = sheet_obj.cell(row = i, column = 1)
	name = sheet_obj.cell(row = i, column = 2)
	party = sheet_obj.cell(row = i, column = 3)
	crime_rec = sheet_obj.cell(row = i, column = 4)
	severity = sheet_obj.cell(row = i, column = 5)
	fam_tree = sheet_obj.cell(row = i, column = 6)
	prop = sheet_obj.cell(row = i, column = 7)
	input_data[sr_no.value] = {'candidate' : name.value,'party':party.value,'criminal_rec': crime_rec.value,
						'severity': severity.value,'family_members': fam_tree.value,'property': prop.value}

print 'Politics Data', input_data

for key in input_data:
	if input_data.get(key)['criminal_rec'] == 'Y' or input_data.get(key)['criminal_rec'] == 'YES':
		if input_data.get(key)['severity'] <= 5:
			if input_data.get(key)['family_members'] <= 3:
				if input_data.get(key)['property'] < 50000:
					success_prob = 80
				else:
					success_prob = 70
			else:
				if input_data.get(key)['property'] < 50000:
					success_prob = 60
				else:
					success_prob = 50
		else:
			if input_data.get(key)['family_members'] <= 3:
				if input_data.get(key)['property'] <= 50000:
					success_prob = 40
				else:
					success_prob = 30
			else:
				if input_data.get(key)['property'] <= 50000:
					success_prob = 20
				else:
					success_prob = 10

	else:
		if input_data.get(key)['criminal_rec'] == 'N' or input_data.get(key)['criminal_rec'] == 'NO':
			if input_data.get(key)['family_members'] <= 3:
				if input_data.get(key)['property'] <= 50000:
					success_prob = 90
				else:
					success_prob = 80	
			else:
				if input_data.get(key)['property'] <= 50000:
					success_prob = 70
				else:
					success_prob = 60

	result_data[key] ={'Name' : input_data.get(key)['candidate'],'Success_Prob' : success_prob} 

	tuple2 = (key,input_data[key]['candidate'],success_prob)
	sheet_obj1.append(tuple2)
	workbook1.save(path1)

#print '\n\nResult Data:', result_data

key_list = []
value_list = []

for k in result_data:
	key_list.append(result_data.get(k)['Name'])
#print '\n\t\tKey List: ', key_list

for v in result_data:
	value_list.append(result_data.get(v)['Success_Prob'])

#print '\n\t\tValue List: ', value_list

series = pd.Series(value_list,key_list).sort_values(ascending = True)
#print '\n\t\tSeries\n',series

series.plot(kind = 'bar',color = 'purple',title = 'Candidate Vs Success %',legend = True)
plt.grid(False)
plt.legend(['Candidates'])
plt.hlines(75, 75,.5, linestyles='dashed')
#plt.xticks(range(0,len(key_list)),key_list)
plt.yticks(range(0,101,20))
plt.ylabel('Success Percentage ')
plt.xlabel('Candidates')
plt.show()

'''
plt.bar(range(0,len(key_list)),value_list)
plt.grid(False)
plt.title('Candidate Vs Success %')
plt.legend(['Candidates'])
plt.ylabel('Success Percentage ')
plt.xlabel('Candidates')
plt.xticks(range(0,len(key_list)),key_list)
plt.yticks(range(0,100,20))
plt.show()

plt.hist(key_list,histtype = 'bar', rwidth = 0.25, color = 'orange')
plt.grid(False)
plt.title('Candidate Vs Success %')
plt.legend(['Candidates'])
plt.ylabel('Success Percentage ')
plt.xlabel('Candidates')
plt.show()
'''